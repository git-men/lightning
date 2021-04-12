import codecs
import csv
from collections import OrderedDict

from django.http import HttpResponse
from pydash import objects
import openpyxl
from openpyxl.styles.alignment import Alignment
from tempfile import NamedTemporaryFile

from api_basebone.core import gmeta
from api_basebone.core.decorators import BSM_ADMIN_COMPUTED_FIELDS_MAP
from api_basebone.utils.gmeta import get_attr_in_gmeta_class
from api_basebone.utils.meta import get_all_relation_fields
from api_basebone.utils.timezone import local_timestamp


def get_no_concrete_or_m2m(model):
    """获取反向或者对对多字段名称列表"""
    return [
        item.name
        for item in get_all_relation_fields(model)
        if not item.concrete or item.many_to_many
    ]


def get_merge_fields(model, serializer_class, export_config):
    """
    针对合并的情况下，返回对应的字段列表和字段名称

    获取指定模型的反向字段和多对多字段列表

    返回的是 {
        field_nane: field_verbose_name
    }
    """
    default_fields = OrderedDict()

    # 指定导出的字段
    export_fields = export_config['fields']

    for item in model._meta.get_fields():
        if export_fields:
            if not item.concrete:
                default_fields[item.name] = item.related_model._meta.verbose_name
            else:
                default_fields[item.name] = item.verbose_name
        else:
            if item.concrete and not item.many_to_many:
                default_fields[item.name] = item.verbose_name

    # 添加 model GMeta 下声明的计算属性字段
    computed_fields = get_attr_in_gmeta_class(model, gmeta.GMETA_COMPUTED_FIELDS, [])
    for field in computed_fields:
        default_fields[field['name']] = field.get('display_name', field['name'])

    # 添加 admin 中声明的计算属性字段
    # 添加 model GMeta 下声明的计算属性字段
    admin_computed_fields = getattr(model, BSM_ADMIN_COMPUTED_FIELDS_MAP, {})
    for key, value in admin_computed_fields.items():
        default_fields[key] = value['display_name']

    if not isinstance(export_fields, (list, tuple)) or not export_fields:
        return default_fields

    fields = OrderedDict()

    # 反向和多对多的字段列表
    no_concrete_or_m2m_fields = get_no_concrete_or_m2m(model)

    for item in export_fields:
        # 针对多对多或者反向做对应的处理
        item_str = item[0] if isinstance(item, (list, tuple)) else item
        item_split = item_str.split('.')
        if item_split[0] in no_concrete_or_m2m_fields:
            fields[item_str] = item_str
            continue

        if isinstance(item, (list, tuple)):
            fields[item[0]] = item[1]
        else:
            fields[item] = default_fields.get(item, item)
    return fields


def row_data_merge(model, fields, data, export_fields):
    """获取不合并的行数据"""
    result = []
    # 反向和多对多的字段列表
    no_concrete_or_m2m_fields = get_no_concrete_or_m2m(model)

    for key in fields.keys():
        if key.split('.')[0] in no_concrete_or_m2m_fields:
            if not export_fields:
                result.append('')
                continue

            inner_data = get_data_from_dict(data, key)
            if not inner_data:
                result.append('')
                continue

            result.append('\n'.join(inner_data))
        else:
            item_data = get_data_from_dict(data, key)
            row_item_data = item_data[0] if item_data else ''
            result.append(row_item_data)
    return [result]


def get_no_merge_fields(model, serializer_class, export_config):
    """
    针对不合并的情况下，返回对应的字段列表和字段名称

    获取指定模型的反向字段和多对多字段列表

    返回的是 {
        field_name: field_verbose_name
    }
    """
    default_fields = OrderedDict()

    # 指定导出的字段
    export_fields = export_config['fields']

    for item in model._meta.get_fields():
        if export_fields:
            if not item.concrete:
                default_fields[item.name] = item.related_model._meta.verbose_name
            else:
                default_fields[item.name] = item.verbose_name
        else:
            if item.concrete and not item.many_to_many:
                default_fields[item.name] = item.verbose_name

    # 添加 model GMeta 下声明的计算属性字段
    computed_fields = get_attr_in_gmeta_class(model, gmeta.GMETA_COMPUTED_FIELDS, [])
    for field in computed_fields:
        default_fields[field['name']] = field.get('display_name', field['name'])

    # 添加 admin 中声明的计算属性字段
    # 添加 model GMeta 下声明的计算属性字段
    admin_computed_fields = getattr(model, BSM_ADMIN_COMPUTED_FIELDS_MAP, {})
    for key, value in admin_computed_fields.items():
        default_fields[key] = value['display_name']

    if not isinstance(export_fields, (list, tuple)) or not export_fields:
        return default_fields

    fields = OrderedDict()

    # 反向和多对多的字段列表
    no_concrete_or_m2m_fields = get_no_concrete_or_m2m(model)

    for item in export_fields:

        if isinstance(item, (list, tuple)):
            fields[item[0]] = item[1]
        else:
            item_split = item.split('.')
            if item_split[0] in no_concrete_or_m2m_fields:
                fields[item] = item
            else:
                fields[item] = default_fields.get(item, item)
    return fields


def row_data_no_merge(model, fields, instance_data, export_fields):
    """不合并行数据的处理"""

    result = []
    max_nest_list = 0

    data = {}
    for key in export_fields:
        if isinstance(key, (list, tuple)):
            key = key[0]
        data[key] = get_data_from_dict(instance_data, key)
        if len(data[key]) > max_nest_list:
            max_nest_list = len(data[key])
    if max_nest_list == 0:
        result.append([objects.get(data, key) for key in fields.keys()])
    else:
        for index in range(0, max_nest_list):
            row_data = []
            for key in fields.keys():
                try:
                    value = data[key][index]
                except Exception:
                    value = ''
                row_data.append(value)
            result.append(row_data)
    return result


def get_data_from_dict(out_data, key):
    """从字典中获取对应的数据

    {
        'age': [
            {"staff": {"u": [{"mm": 45}]}},
            {"staff": {"u": [{"mm": 46}]}}
        ]
    }

    key 为 'age.staff.u.mm' 时

    处理后返回 [45, 46]
    """

    result = []

    key_split = key.split('.')

    def inner_hand(data, index):
        key = key_split[index]
        value = objects.get(data, key)

        if index == (len(key_split) - 1):
            if value is not None:
                result.append(str(value))
            return

        if isinstance(value, list):
            k_index = index + 1
            for item in value:
                inner_hand(item, k_index)
        else:
            inner_hand(value, index + 1)

    inner_hand(out_data, 0)

    return result


def csv_render(model, queryset, serializer_class, export_config=None):
    """渲染数据"""
    app_label, model_name = model._meta.app_label, model._meta.model_name
    file_name = f'{app_label}-{model_name}-{local_timestamp()}'

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{file_name}.csv"'

    response.write(codecs.BOM_UTF8)
    writer = csv.writer(response)

    if export_config['merge_bref']:
        fields = get_merge_fields(model, serializer_class, export_config)
        verbose_names = fields.values()

        writer.writerow(verbose_names)

        # 处理结果集
        queryset_iter = queryset if isinstance(queryset, list) else queryset.all()

        for instance in queryset_iter:
            instance_data = serializer_class(instance).data

            # 写入一行数据
            writer.writerow(
                row_data_merge(model, fields, instance_data, export_config['fields'])
            )
    else:
        # 不合并的场景
        fields = get_no_merge_fields(model, serializer_class, export_config)
        verbose_names = fields.values()
        writer.writerow(verbose_names)

        # 处理结果集
        queryset_iter = queryset if isinstance(queryset, list) else queryset.all()

        for instance in queryset_iter:
            instance_data = serializer_class(instance).data
            # 写入一行数据
            rows = row_data_no_merge(model, fields, instance_data, export_config['fields'])            
            writer.writerows(rows)
    return response


def excel_render(model, queryset, serializer_class, export_config=None):
    app_label, model_name = model._meta.app_label, model._meta.model_name
    file_name = f'{app_label}-{model_name}-{local_timestamp()}'

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{file_name}.xlsx"'

    datas = []
    
    get_fields = get_merge_fields if export_config['merge_bref'] else get_no_merge_fields
    row_data = row_data_merge if export_config['merge_bref'] else row_data_no_merge

    fields = get_fields(model, serializer_class, export_config)
    # 处理结果集
    queryset_iter = queryset if isinstance(queryset, list) else queryset.all()
    for instance in queryset_iter:
        instance_data = serializer_class(instance).data
        rows = row_data(model, fields, instance_data, export_config['fields'])
        datas.append(rows)
    
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    titles = list(fields.values())
    print('titles: ', titles)
    for i in range(len(titles)):
        sheet.cell(1, i + 1, titles[i])
    
    insert_line = 2
    for rows in datas:
        for row in rows:
            for i in range(len(row)):
                cell = sheet.cell(insert_line, i + 1, row[i])
                cell.alignment = Alignment(wrap_text=True)
            insert_line += 1
    with NamedTemporaryFile() as tmp:
        workbook.save(tmp.name)
        tmp.seek(0)
        content = tmp.read()
        response.write(content)
    return response

def render(model, queryset, serializer_class, export_config=None, file_type='excel'):        
    if file_type == 'csv':
        return csv_render(model, queryset, serializer_class, export_config)
    return excel_render(model, queryset, serializer_class, export_config)