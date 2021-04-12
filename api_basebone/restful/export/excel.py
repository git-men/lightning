import datetime
import openpyxl
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
import collections
import requests
from time import time
from tempfile import NamedTemporaryFile
from django.db.models.fields.related import ManyToManyField
from django.db.models.fields.reverse_related import ManyToManyRel, ManyToOneRel
from django.db.models.query import QuerySet
from django.http import HttpResponse
from django.db import models, transaction

from rest_framework.exceptions import ValidationError

from api_basebone.utils.timezone import local_timestamp
from api_basebone.restful.export.formatter import format
from api_basebone.restful.forms import get_form_class
from api_basebone.drf.response import success_response

def get_attribute(instance, field_path, formatter=None):
    # 获取对象属性值,并对其进行格式化
    print(f'getting {field_path} from instance: {instance}')
    def _get(obj, field):
        rs = getattr(obj, field, None)
        # 如果是一对多，多对多的refManager，把它all一下。
        opts = obj.__class__._meta
        try:
            field_obj = opts.get_field(field)
        except:  # FIXME 找不到该字段是，可能是计算字段。
            field_obj = None
        if field_obj.__class__ in [ManyToManyField, ManyToManyRel, ManyToOneRel]:
            return rs.all()
        if getattr(field_obj, 'choices', None):
            choices = dict(field_obj.choices)
            if rs is None and rs not in choices:
                return ''
            return dict(field_obj.choices)[rs]
        return rs

    # 逐层获取属性
    result = instance
    paths = field_path.split('.')
    for path in paths:

        # 这里的result可以是Queryset，可以是List
        if isinstance(result, QuerySet) or isinstance(result, list):
            inter = []
            for item in result:
                rs = _get(item, path)
                if isinstance(rs, QuerySet) or isinstance(rs, list):
                    inter += rs
                else:
                    inter.append(rs)
            result = inter
        else:
            result = _get(result, path)
    
    # 格式化
    if not formatter:
        return result
    if isinstance(result, QuerySet) or isinstance(result, list):
        return [format(rs, formatter['type'], formatter['params']) for rs in result]
    return format(result, formatter['type'], formatter['params'])

def convertToTitle(n):
    return ('' if n <= 26 else convertToTitle((n-1)//26)) + chr((n-1)%26+ord('A'))

def export_excel(config, queryset, detail=None):
    """
    导出Excel
    - config: 导出配置
    - queryset: 导出结果集
    - detail: 关联详情数据

    config结构示例：
    {
        "version": "v3",
        "template": "https://xxxx.com/yyy.xlsx",
        "list_start_line": 8,
        "detail_mapping": [{
            "position": "A1",
            "field": "superior.name",
            "formatter": {}
        }],
        "list_mapping": [{
            "column": "A",
            "field": "a", 
            "formatter": {
                "type": "prefix", 
                "params": {
                    "prefix": "姓  名：",
                }
            }, 
        }]
    }

    """
    app_label, model_name = queryset.model._meta.app_label, queryset.model._meta.model_name
    file_name = f'{app_label}-{model_name}-{local_timestamp()}'

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{file_name}.xlsx"'

    if not isinstance(config, dict):
        # 如果不是配置，可能是Key，通过其他方法获取配置内容。
        pass
    use_template = "template" in config
    if use_template:
        # 有模板，通过模板文件的md5值，从临时目录中获取一下。
        data = requests.get(config['template']).content
        with NamedTemporaryFile() as tmp:
            with open(tmp.name + '.xlsx', 'wb') as f:
                f.write(data)
            print('下载的文件：', tmp.name)

        workbook = openpyxl.load_workbook(tmp.name + '.xlsx')
        sheet = workbook.worksheets[0]
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active

    detail_mapping = config.get('detail_mapping', [])
    
    list_mapping = config.get('list_mapping', [])

    model_fields = dict([(f.name, f) for f in queryset.model._meta.fields])
    print('model_fields', model_fields)
    # 渲染列表
    if list_mapping:
        if use_template:  # 指定模板
            row = config.get('list_start_line', 1)
            end = config.get('list_end_line')
            for instance in queryset:
                for field in list_mapping:
                    sheet[f'{field["column"]}{row}'] = get_attribute(instance, field['field'], field.get('formatter', None))
                row += 1

            sheet.delete_rows(row, end - row + 1)
        else:
            row = len(detail_mapping) + 1
            
            # 输出标題
            col = 1
            for field in list_mapping:
                title = field.get('title', '') # FIXME 考虑没有指定Title的情况。
                sheet.cell(row, col, title)
                field_name = field.get('field')  # FIXME 考虑嵌套的情况, 如果嵌套，结果只是没校验，但输出没问題。
                if field_name in model_fields:
                    if model_fields[field_name].choices:
                        # 生成Validation
                        formula = ','.join([v[1] for v in model_fields[field_name].choices])
                        dv = DataValidation(type="list", formula1=f'"{formula}"', allow_blank=True)
                        dv.error = '单元格的值不在有效范围'
                        dv.errorTitle = '错误值'
                        # dv.prompt = '选择以下选项中的值'
                        # dv.promptTitle = '请选择'
                        sheet.add_data_validation(dv)
                        title = convertToTitle(col)
                        dv.add(f'{title}{row + 1}:{title}1048576')
                        print('校验范围： ',f'{title}{row + 1}:{title}1048576', formula)
                col += 1
            row += 1

            for instance in queryset:
                col = 1
                for field in list_mapping:
                    value = get_attribute(instance, field['field'], field.get('formatter', None))
                    sheet.cell(row, col, value)
                    col += 1
                row += 1
    
    if detail_mapping:
        # 渲染关联详情
        if use_template:  # 模型指定位置
            qs_len = len(queryset)
            for dfield in detail_mapping:
                position = dfield['position']
                if isinstance(position, str):
                    sheet[dfield['position']] = get_attribute(detail, dfield['field'], dfield.get('formatter', None))
                elif isinstance(position, dict):
                    row = position['related_row'] + config.get('list_start_line') + qs_len - 1
                    sheet.cell(row, position['column'], get_attribute(detail, dfield['field'], dfield.get('formatter', None)))
        else:  # 自动生成，从第一行起，合并单元格。
            cur = 1
            col_cnt = len(config.get('list_mapping', []))
            for dfield in detail_mapping:
                sheet.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=col_cnt)
                sheet[f'A{cur}'] = get_attribute(detail, dfield['field'], dfield.get('formatter', None))
                cur += 1

    with NamedTemporaryFile() as tmp:
        workbook.save(tmp.name)
        tmp.seek(0)
        content = tmp.read()
        response.write(content)
    return response


def import_excel(config, content, queryset, request, detail=None):
    """导入Exce
    参数：
    1. config: 导入配置
    2. content: 文件内容
    3. detail: 上层数据

    config的结构：
    {	
        "type": "create",  // create | update
        "template": "",  // 可用模板，可不用模板。
        "update_by": ["user.first_name", "user.last_name"],  // 更新的话需要指定条件
        "list_mapping": [  // 字段映射，暂不允许关联更新
            {"column": "", "field": ""},
        ],
        "list_start_line": 10  // 数据开始行
    }
    """
    with NamedTemporaryFile(suffix='.xlsx') as tmp:
        tmp.write(content)
        workbook = load_workbook(tmp.name)
    print('excel file name: ', tmp.name)

    sheet = workbook.worksheets[0]

    list_mapping = config['list_mapping']
    start_line = config.get('list_start_line', 1)
    line = start_line

    model_fields = dict([(f.name, f) for f in queryset.model._meta.fields])

    data = []
    eof = False
    while not eof:
        row_data = {}
        for field in list_mapping:
            value = sheet[f'{field["column"]}{line}'].value  # TODO 考虑多层级场景
            if isinstance(value, datetime.datetime) and isinstance(model_fields[field["field"]], models.DateField):
                value = value.date()
            if field["field"] in model_fields and model_fields[field["field"]].choices:
                choices = dict([(c[1], c[0]) for c in model_fields[field["field"]].choices])
                if value in choices:
                    value = choices[value]
            row_data[field["field"]] = value
        line += 1
        if not [val for val in row_data.values() if val]:
            eof = True
        else:
            data.append(row_data)

    if config['type'] == 'create':
        serializer = get_form_class(queryset.model, 'create', request=request)(data=data, many=True)
    elif config['type'] == 'update':
        # 更新，是有条件的更新
        # update_by = config.get('update_by', [])
        pk_name = queryset.model._meta.pk.name
        pk_values = [row[pk_name] for row in data]
        condition = {f'{pk_name}__in': pk_values}
        print('condition: ', condition)
        instances = queryset.filter(**condition)
        
        # FIXME 如果data里面有instances里面没有的数据时，要报无权修改
        serializer = get_form_class(queryset.model, 'update', request=request, batch=True)(instances, data=data, partial=True, many=True)
    
    if not serializer.is_valid(raise_exception=False):
        error_details = []
        errors = serializer.errors
        for idx in range(len(errors)):
            if errors[idx]:
                error_details.append({
                    "line": start_line + idx,
                    "error": errors[idx]
                })
        raise ValidationError(error_details)
    with transaction.atomic():
        serializer.save()
    return success_response()
